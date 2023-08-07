from openpyxl import Workbook
from openpyxl import load_workbook


class ExcelWork:
    def __init__(self, date):
        self.wb = None
        self.ws = None
        self.date = date

    def create_work_book(self, file_path):
        self.wb = Workbook()
        self.wb['Sheet'].title = self.date
        self.ws = self.wb.active
        self.ws.append(['S.No', 'Date', 'Full Name', 'Emp ID', 'Project/Platform/Team Name', 'Office Location', 'Reporting Manager Name'])
        self.wb.save(file_path)

    def load_work_book(self, file_path):
        self.wb = load_workbook(file_path)
        self.ws = self.wb.active

    def write_data(self, emp_data, file_path):
        self.ws.append(emp_data)
        self.wb.save(file_path)

    def get_row_count(self):
        return self.ws.max_row


