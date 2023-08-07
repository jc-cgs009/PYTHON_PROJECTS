from openpyxl import load_workbook


def get_row_count(file, sheet_name):
    wb = load_workbook(file)
    ws = wb[sheet_name]
    return ws.max_row


def get_column_count(file, sheet_name):
    wb = load_workbook(file)
    ws = wb[sheet_name]
    return ws.max_column


def read_excel_data(file, sheet_name, row_num, col_num):
    wb = load_workbook(file)
    ws = wb[sheet_name]
    return ws.cell(row=row_num, column=col_num).value


def write_excel_data(file, sheet_name, row_num, col_num, val):
    wb = load_workbook(file)
    ws = wb[sheet_name]
    ws.cell(row=row_num, column=col_num).value = val
    wb.save(file)
